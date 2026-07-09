#!/usr/bin/env python3
"""
update_catalogo_maestro.py
Sincroniza el dashboard-maestro con los datos reales del Google Sheet "Nueva Sede Arqueo Bar".
Actualiza catalogo.json con totales de ventas bar correctos para cada mes.
"""

import json
import os
import re
from datetime import datetime
import requests
import openpyxl

SHEET_ID = "1pYhvrTfPAVzxW8qeaR8LU8pl6YhEUP81"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOGO_PATH = os.path.join(SCRIPT_DIR, "catalogo.json")

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

MES_NAMES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def clean_money(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace("$", "").replace(".", "").replace(",", ".").strip()
        if cleaned in ("", "-"):
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


def fix_date(dt, expected_month, expected_year):
    if not isinstance(dt, datetime):
        return dt
    if dt.year == expected_year:
        return dt
    if dt.month == expected_month and dt.day == 1:
        real_day = dt.year - 2000
        if 1 <= real_day <= 31:
            return datetime(expected_year, expected_month, real_day)
    return dt


def parse_sheet_name(tab_name):
    m = re.match(r"([a-záéíóúñ]+).*?(\d+)", tab_name, re.IGNORECASE)
    if m:
        month = MESES.get(m.group(1).lower())
        year = 2000 + int(re.sub(r"[^0-9]", "", m.group(2)))
        return month, year
    return None, None


def extract_sheet_data(ws, tab_name):
    """Extrae datos de una pestaña: totales y métricas."""
    raw_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9, values_only=True):
        raw_data.append(row)

    month, year = parse_sheet_name(tab_name)
    if month is None:
        return None

    # Encontrar fila TOTAL y contar días con datos
    total_alegra = 0
    dias_con_dato = 0
    meta_mes = None

    for row in raw_data[2:]:
        fecha_val = row[0] if len(row) > 0 else None

        # Capturar TOTAL row
        if isinstance(fecha_val, str) and "total" in fecha_val.lower():
            total_alegra = clean_money(row[1]) if len(row) > 1 else 0
            continue

        # Capturar META MES
        if isinstance(fecha_val, str) and "meta" in fecha_val.lower():
            meta_val = row[1] if len(row) > 1 else None
            if meta_val:
                if isinstance(meta_val, str):
                    meta_val = clean_money(meta_val)
                elif isinstance(meta_val, (int, float)):
                    meta_val = int(meta_val)
                meta_mes = meta_val if meta_val > 0 else None
            continue

        # Saltar filas vacías o no-fecha
        if fecha_val is None:
            continue
        if isinstance(fecha_val, str):
            continue

        fecha_corrected = fix_date(fecha_val, month, year)
        cierreAlegra = clean_money(row[1]) if len(row) > 1 else 0
        cierreFormato = clean_money(row[2]) if len(row) > 2 else 0

        if cierreAlegra > 0 or cierreFormato > 0:
            dias_con_dato += 1

    if total_alegra == 0 or dias_con_dato == 0:
        return None

    # Calcular promedio y crecimiento
    promedio_diario = total_alegra // dias_con_dato

    return {
        "total_ventas": total_alegra,
        "dias_con_dato": dias_con_dato,
        "promedio_diario": promedio_diario,
        "meta_mes": meta_mes,
    }


def main():
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Sincronizando dashboard-maestro...")

    # 1. Descargar XLSX
    resp = requests.get(XLSX_URL, allow_redirects=True, timeout=30)
    if resp.status_code != 200:
        print(f"  ERROR descargando sheet: {resp.status_code}")
        return 1

    xlsx_path = "/tmp/arqueo_bar_maestro.xlsx"
    with open(xlsx_path, "wb") as f:
        f.write(resp.content)

    # 2. Extraer datos de todas las pestañas
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_data = {}
    for tab_name in wb.sheetnames:
        month, year = parse_sheet_name(tab_name)
        if month is None:
            continue
        data = extract_sheet_data(wb[tab_name], tab_name)
        if data:
            mes_key = f"{MES_NAMES[month]} {year}"
            sheet_data[mes_key] = data
            print(f"  {mes_key}: ${data['total_ventas']:,} ({data['dias_con_dato']} días, meta={data.get('meta_mes','--')})")

    # 3. Leer y actualizar catalogo.json
    if not os.path.exists(CATALOGO_PATH):
        print("  ERROR: catalogo.json no encontrado")
        return 1

    with open(CATALOGO_PATH, "r", encoding="utf-8") as f:
        catalogo = json.load(f)

    catalogo["actualizado"] = datetime.now().strftime("%Y-%m-%d")

    # Reconstruir periodos con datos actualizados
    nuevos_periodos = []

    for i, p in enumerate(catalogo.get("periodos", [])):
        mes = p["mes"]
        if mes in sheet_data:
            sd = sheet_data[mes]

            # Calcular crecimiento respecto al mes anterior (usando datos NUEVOS)
            crecimiento = None
            if i > 0 and catalogo["periodos"][i - 1]["mes"] in sheet_data:
                prev_mes = catalogo["periodos"][i - 1]["mes"]
                prev_total = sheet_data[prev_mes]["total_ventas"]
                if prev_total > 0:
                    crecimiento = round((sd["total_ventas"] - prev_total) / prev_total * 100, 1)

            ventas_bar = {
                "total_ventas": sd["total_ventas"],
                "dias_con_dato": sd["dias_con_dato"],
                "promedio_diario": sd["promedio_diario"],
            }
            if crecimiento is not None:
                ventas_bar["crecimiento"] = crecimiento
            if sd.get("meta_mes"):
                ventas_bar["meta_mes"] = sd["meta_mes"]

            p["ventas_bar"] = ventas_bar

        nuevos_periodos.append(p)

    # Agregar meses nuevos del sheet que no están en catalogo
    existing_meses = {p["mes"] for p in catalogo.get("periodos", [])}
    for mes_key, sd in sheet_data.items():
        if mes_key not in existing_meses:
            nuevos_periodos.append({
                "mes": mes_key,
                "ventas_bar": {
                    "total_ventas": sd["total_ventas"],
                    "dias_con_dato": sd["dias_con_dato"],
                    "promedio_diario": sd["promedio_diario"],
                },
                "inventario": None,
            })
            print(f"  + NUEVO periodo: {mes_key}")

    catalogo["periodos"] = nuevos_periodos

    # 4. Guardar
    with open(CATALOGO_PATH, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)

    print(f"  ✅ catalogo.json actualizado ({len(catalogo['periodos'])} periodos)")
    return 0


if __name__ == "__main__":
    exit(main())
